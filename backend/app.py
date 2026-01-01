from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import pandas as pd
import math
import tempfile
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import os

app = Flask(__name__)
CORS(app)

@app.route("/")
def home():
    return "Backend online"

@app.route("/gerar_excel", methods=["POST"])
def gerar_excel():
    data = request.json

    dados = data["dados"]
    motoboys = data["motoboys"]
    entregas = data["entregas"]

    linhas_m, linhas_c = [], []
    total_motoboys, total_clientes = 0, 0

    for i, m in enumerate(motoboys):
        ent = [e for e in entregas if e["id_moto"] == i]

        qtd = len(ent)
        km_exc_m, km_exc_c = 0, 0

        for e in ent:
            km = float(e["km"])
            km_exc_m += max(0, math.floor(km - dados["kmm"]))
            km_exc_c += max(0, math.floor(km - dados["kmc"]))

        taxa_m = qtd * dados["tm"] + km_exc_m
        taxa_c = qtd * dados["tc"] + km_exc_c

        diaria_m = dados["dcoord"] if m["tipo"] == "Coordenação" else dados["dm"]
        diaria_c = dados["dc"]

        chuva_m = dados["chuva_moto"] if m["aplica_chuva"] and qtd > 0 else 0
        chuva_c = dados["chuva_cliente"] if m["aplica_chuva"] and qtd > 0 else 0

        total_m = diaria_m + taxa_m + chuva_m - m["vale"]
        total_c = diaria_c + taxa_c + chuva_c

        total_motoboys += total_m
        total_clientes += total_c

        linhas_m.append([
            m["nome"], m["tipo"], qtd, km_exc_m,
            taxa_m, diaria_m, chuva_m, m["vale"], total_m
        ])

        linhas_c.append([
            m["nome"], qtd, km_exc_c,
            taxa_c, diaria_c, chuva_c, total_c
        ])

    df_m = pd.DataFrame(linhas_m, columns=[
        "Motoboy","Tipo","Qtd","KM Exc",
        "Taxas","Diária","Chuva","Vale","TOTAL"
    ])

    df_c = pd.DataFrame(linhas_c, columns=[
        "Motoboy","Qtd","KM Exc",
        "Taxas","Diária","Chuva","TOTAL"
    ])

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.close()

    with pd.ExcelWriter(tmp.name, engine="openpyxl") as writer:
        df_m.to_excel(writer, sheet_name="Motoboys", index=False)
        df_c.to_excel(writer, sheet_name="Cliente", index=False)

    wb = load_workbook(tmp.name)
    borda = Border(*(Side(style="thin"),)*4)

    for ws in wb.worksheets:
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18
            for c in col:
                c.alignment = Alignment(horizontal="center")
                c.border = borda
        for c in ws[1]:
            c.font = Font(bold=True)

    wb.save(tmp.name)

    return send_file(
        tmp.name,
        as_attachment=True,
        download_name="relatorio_logistica.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port)
