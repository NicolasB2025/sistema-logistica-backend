from flask import Flask, request, send_file
from flask_cors import CORS
import pandas as pd
import math
import tempfile
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

app = Flask(__name__)
CORS(app)

@app.route("/gerar_excel", methods=["POST"])
def gerar_excel():
    data = request.json
    dados = data["dados"]
    motoboys = data["motoboys"]
    entregas = data["entregas"]

    linhas_m, linhas_c = [], []
    total_motoboys, total_clientes = 0, 0

    for i, m in enumerate(motoboys):
        ent = [e for e in entregas if e["id_moto"] == i]

        qtd_entregas = len(ent)
        soma_km_exc_moto = 0
        soma_km_exc_cliente = 0

        for e in ent:
            km_real = float(e["km"])

            # 🔴 REGRA FINAL: só conta km COMPLETO acima do limite
            exc_moto = math.floor(km_real - dados["kmm"])
            exc_cliente = math.floor(km_real - dados["kmc"])

            if exc_moto > 0:
                soma_km_exc_moto += exc_moto
            if exc_cliente > 0:
                soma_km_exc_cliente += exc_cliente

        # TAXAS
        taxa_motoboy = (qtd_entregas * dados["tm"]) + soma_km_exc_moto
        taxa_cliente = (qtd_entregas * dados["tc"]) + soma_km_exc_cliente

        # DIÁRIAS
        diaria_moto = dados["dcoord"] if m["tipo"] == "Coordenação" else dados["dm"]
        diaria_cliente = dados["dc"]

        # CHUVA
        aplica_chuva = m["aplica_chuva"] and qtd_entregas > 0
        chuva_moto = dados["chuva_moto"] if aplica_chuva else 0
        chuva_cliente = dados["chuva_cliente"] if aplica_chuva else 0

        total_pagar_moto = diaria_moto + taxa_motoboy + chuva_moto - m["vale"]
        total_cobrar_cliente = diaria_cliente + taxa_cliente + chuva_cliente

        total_motoboys += total_pagar_moto
        total_clientes += total_cobrar_cliente

        linhas_m.append([
            m["nome"], m["tipo"], qtd_entregas, soma_km_exc_moto,
            taxa_motoboy, diaria_moto, chuva_moto, m["vale"], total_pagar_moto
        ])

        linhas_c.append([
            m["nome"], qtd_entregas, soma_km_exc_cliente,
            taxa_cliente, diaria_cliente, chuva_cliente, total_cobrar_cliente
        ])

    df_m = pd.DataFrame(linhas_m, columns=[
        "Motoboy","Tipo","Qtd Entregas","KM Excedentes",
        "Taxas Motoboy (R$)","Diária (R$)",
        "Chuva Motoboy (R$)","Vale (R$)","TOTAL A PAGAR (R$)"
    ])

    df_c = pd.DataFrame(linhas_c, columns=[
        "Motoboy","Qtd Entregas","KM Excedentes",
        "Taxas Cliente (R$)","Diária Cliente (R$)",
        "Chuva Cliente (R$)","TOTAL CLIENTE (R$)"
    ])

    df_m.loc[len(df_m)] = ["TOTAL GERAL","","","","","","","", total_motoboys]
    df_c.loc[len(df_c)] = ["TOTAL GERAL","","","","","", total_clientes]

    resumo_dia = pd.DataFrame({
        "Campo": [
            "Cliente","Período","Localização",
            "Diária Motoboy","Diária Coordenação",
            "Taxa Mínima Motoboy","KM Limite Motoboy",
            "Diária Cliente","Taxa Mínima Cliente","KM Limite Cliente",
            "Taxa Chuva Motoboy","Taxa Chuva Cliente","Custo Operacional"
        ],
        "Valor": [
            dados["cliente"], dados["periodo"], dados["local"],
            dados["dm"], dados["dcoord"],
            dados["tm"], dados["kmm"],
            dados["dc"], dados["tc"], dados["kmc"],
            dados["chuva_moto"], dados["chuva_cliente"], dados["custo"]
        ]
    })

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.close()

    with pd.ExcelWriter(tmp.name, engine="openpyxl") as writer:
        resumo_dia.to_excel(writer, sheet_name="Resumo_Dia", index=False)
        df_m.to_excel(writer, sheet_name="Resumo_Motoboys", index=False)
        df_c.to_excel(writer, sheet_name="Resumo_Cliente", index=False)

    wb = load_workbook(tmp.name)
    borda = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            for cell in col:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borda
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col[0].column_letter].width = max_len + 4
        for cell in ws[1]:
            cell.font = Font(bold=True)

    wb.save(tmp.name)

    return send_file(
        tmp.name,
        as_attachment=True,
        download_name="relatorio_logistica_final.xlsx"
    )

if __name__ == "__main__":
    app.run(debug=True)

